"""
ESG Risk Report Generator
=========================
Script tạo báo cáo Excel từ file esg_data.json

Cách sử dụng:
    python create_esg_report.py

File output: esg_risk_report_[timestamp].xlsx
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os

# Đọc data từ file JSON
DATA_FILE = "esg_data.json"
OUTPUT_DIR = "D:/ESG"

def load_data():
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def create_excel(events):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi tiet ESG"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Cong ty", "Loai", "Ngay", "Tom tat su viec", "Muc do", "Nguon", "Link bai bao"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Color by type
    type_colors = {
        "E": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "S": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "G": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    }

    # Data rows
    for row, event in enumerate(events, 2):
        fill = type_colors.get(event["type"], None)
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
            if fill:
                ws.cell(row=row, column=col).fill = fill
        
        ws.cell(row=row, column=1, value=event["company"])
        ws.cell(row=row, column=2, value=event["type"]).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=event["date"])
        ws.cell(row=row, column=4, value=event["summary"])
        ws.cell(row=row, column=5, value=event["severity"])
        ws.cell(row=row, column=6, value=event["source"])
        
        # Hyperlink
        link_cell = ws.cell(row=row, column=7, value="Xem bai bao")
        link_cell.hyperlink = event["url"]
        link_cell.font = Font(color="0000FF", underline="single")

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 75
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15

    # Summary sheet
    ws2 = wb.create_sheet("Tong hop")
    summary_headers = ["Loai rui ro", "So luong", "Ghi chu"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Count by type
    e_count = sum(1 for e in events if e["type"] == "E")
    s_count = sum(1 for e in events if e["type"] == "S")
    g_count = sum(1 for e in events if e["type"] == "G")
    
    data = [
        ("E - Moi truong", e_count, ""),
        ("S - Xa hoi", s_count, ""),
        ("G - Quan tri", g_count, ""),
        ("TONG", len(events), "")
    ]
    for row, (loai, sl, ghichu) in enumerate(data, 2):
        ws2.cell(row=row, column=1, value=loai).border = border
        ws2.cell(row=row, column=2, value=sl).border = border
        ws2.cell(row=row, column=3, value=ghichu).border = border
        if loai == "TONG":
            ws2.cell(row=row, column=1).font = Font(bold=True)
            ws2.cell(row=row, column=2).font = Font(bold=True)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 35

    return wb

def main():
    # Load data
    events = load_data()
    print(f"Loaded {len(events)} events from {DATA_FILE}")
    
    # Create Excel
    wb = create_excel(events)
    
    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{OUTPUT_DIR}/esg_risk_report_{timestamp}.xlsx"
    wb.save(filename)
    
    print(f"Created: {filename}")
    print(f"Total events: {len(events)}")

if __name__ == "__main__":
    main()
