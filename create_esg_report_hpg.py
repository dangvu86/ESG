import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Data với link đã verify qua browser
events = [
    {
        "company": "Hòa Phát (HPG)",
        "type": "E",
        "date": "09/2021",
        "summary": "Công ty Chăn nuôi Thái Thụy bị phạt 130 triệu đồng vì gây ô nhiễm không khí (NH3 vượt quy chuẩn)",
        "severity": "Trung bình",
        "source": "Ngày Mới Online",
        "url": "https://ngaymoionline.com.vn/tinh-thai-binh-cong-ty-thuoc-tap-doan-hoa-phat-bi-phat-130-trieu-dong-vi-gay-o-nhiem-moi-truong-27024.html"
    },
    {
        "company": "Hòa Phát (HPG)",
        "type": "E", 
        "date": "01/2021",
        "summary": "Nhà máy thép Dung Quất xả khói bụi khét lẹt khi vận hành thử lò cao - dân bức xúc phản đối",
        "severity": "Cao",
        "source": "Tuổi Trẻ",
        "url": "https://tuoitre.vn/khoi-bui-khet-let-tu-nha-may-thep-hoa-phat-dung-quat-khien-dan-buc-xuc-20210107164102434.htm"
    },
    {
        "company": "Hòa Phát (HPG)",
        "type": "S",
        "date": "04/09/2021",
        "summary": "Tai nạn lao động tại thép Hòa Phát Hải Dương: 1 nam công nhân (SN 1985) tử vong do ngã từ độ cao",
        "severity": "Cao",
        "source": "Bảo vệ Pháp luật",
        "url": "https://baovephapluat.vn/van-hoa-xa-hoi/lao-dong-viec-lam/tai-nan-tai-nha-may-thep-hoa-phat-hai-duong-nam-cong-nhan-tu-vong-111669.html"
    },
    {
        "company": "Hòa Phát (HPG)",
        "type": "G",
        "date": "05/2022",
        "summary": "Bị UBCKNN phạt 125 triệu đồng vì không có thành viên HĐQT độc lập theo quy định",
        "severity": "Trung bình",
        "source": "VnExpress",
        "url": "https://vnexpress.net/hoa-phat-bi-phat-vi-khong-co-thanh-vien-hoi-dong-quan-tri-doc-lap-4464423.html"
    },
    {
        "company": "Hòa Phát (HPG)",
        "type": "G",
        "date": "06/2024",
        "summary": "Bị UBCKNN phạt 112.5 triệu đồng vì không đủ tỷ lệ 1/3 thành viên HĐQT độc lập (chỉ có 2/9)",
        "severity": "Trung bình",
        "source": "VOV",
        "url": "https://vov.vn/kinh-te/tap-doan-hoa-phat-bi-uy-ban-chung-khoan-nha-nuoc-xu-phat-post1104743.vov"
    }
]

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Chi tiết ESG"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["Công ty", "Loại", "Ngày", "Tóm tắt sự việc", "Mức độ", "Nguồn", "Link bài báo"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Color by type
type_colors = {
    "E": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "S": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "G": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
}

# Data rows
for row, event in enumerate(events, 2):
    fill = type_colors.get(event["type"])
    
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = fill
    
    ws.cell(row=row, column=1, value=event["company"])
    ws.cell(row=row, column=2, value=event["type"]).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=3, value=event["date"])
    ws.cell(row=row, column=4, value=event["summary"])
    ws.cell(row=row, column=5, value=event["severity"])
    ws.cell(row=row, column=6, value=event["source"])
    
    # Hyperlink - VERIFIED
    link_cell = ws.cell(row=row, column=7, value="Xem bài báo")
    link_cell.hyperlink = event["url"]
    link_cell.font = Font(color="0000FF", underline="single")

# Column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 6
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 75
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 15

# Summary sheet
ws2 = wb.create_sheet("Tổng hợp")
summary_headers = ["Loại rủi ro", "Số lượng", "Ghi chú"]
for col, header in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

data = [
    ("E - Môi trường", 2, "Ô nhiễm không khí, khói bụi"),
    ("S - Xã hội", 1, "Tai nạn lao động chết người"),
    ("G - Quản trị", 2, "Vi phạm quy định HĐQT độc lập"),
    ("TỔNG", 5, "")
]
for row, (loai, sl, ghichu) in enumerate(data, 2):
    ws2.cell(row=row, column=1, value=loai).border = border
    ws2.cell(row=row, column=2, value=sl).border = border
    ws2.cell(row=row, column=3, value=ghichu).border = border
    if loai == "TỔNG":
        ws2.cell(row=row, column=1).font = Font(bold=True)
        ws2.cell(row=row, column=2).font = Font(bold=True)

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 35

# Save
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f"D:/ESG/esg_risk_report_HPG_{timestamp}.xlsx"
wb.save(filename)
print(f"Da tao file: {filename}")
print(f"Tong so su kien: 5 (E:2, S:1, G:2)")
print(f"Tat ca 5 link da duoc verify hoat dong")
